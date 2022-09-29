#===================================
# -------  to be improved ----------
# 1- I removed the code below and kept a long protocol with lots of sheets, because it does not copy the charts, so its easiert to just fill and then remove the ones not used
# # create as much sheets as needed
# 2 - for nn in range(len(qlist)):
#     source = xl['(0)']
#     target = xl.copy_worksheet(source)
# 3- En el archivo de protocolo hay que modificar la inclusión del nombre de archivo de mea, set y config
# 4- Incluir un método para que tome automáticamente si la medición se hace en aire o en el baño y corrija el valor de la resistencia (probablemente dos protocolos son necesarios para esto)
# 5- Modificar celda P10 del protocolo a: =IFS(NUMBERVALUE(C23)<30,30,NUMBERVALUE(C23)<100,300,NUMBERVALUE(C23)<300,3000) De forma tal que reconozca la relación del RE automáticamente.
# 5- Tengo que incluir la corrección de drift en el protocolo
# 6- Incluir medición de temperatura
#===================================

import pandas as pd
import io
import os
import sys
import traceback  # to indicate errors and exceptions 
import re
from inspect import currentframe, getframeinfo
import openpyxl as oxl
from openpyxl.formula.translate import Translator
print('package import finished')


def openFile(path2file):
    # open the file, returns a list of the lines of path2file. It is assumed that path2file is a .mea file (text based)
    try:
        with open(path2file, mode='r') as f:  # to ensure it closes correclty
            lines = f.readlines()
            return lines
    except Exception as exception:
        print("There was a problem opening {}".format(path2file))
        print('error in line {}'.format(getframeinfo(currentframe()).lineno))
        print(exception)
        raise

def read_mea_file(lines_of_mea_file, ln_preface = 81):
    '''
    The function reads and arrange the data list of the file called by openFile.
    Inputs:
        lines_of_mea_file: list of str lines read by openFile()
        ln_preface(optional, default=  81): number of lines in the preface of the measurments file. They appear in all sets of measurements.
    Outputs:
        qlistInit: list containing the set measurement prefaces, eg. qlistInit[2]= preface of measurement 3
        qlist: list containing the set measurements, eg. qlist[2]= measurement set number 3
        qlistFinal: list containing the result resume of the measurement set, eg. qlistFinal[2]= resume of measurement set 3
    To process it is preferable to have all measurements together in the same file.
    Note that it doesn´t matter if the measurements have differet number of points.
    '''
    chunks = []
    qlist = []
    chunksInit =[]
    qlistInit = []
    chunksFinal = []
    qlistFinal = []
    nn = 0
    counter = 0
    ii = 0
    for line in lines:
        if (nn>ln_preface and line != '\n' and ii == 0):
            chunks.append(line)
            nn += 1
        elif (nn<ln_preface):
            chunksInit.append(line)
            nn +=1
        elif (nn>ln_preface and line == '\n'):
            ii = 1
            nn += 1
        elif (nn>ln_preface and ii == 1 and line != '***\n'):
            chunksFinal.append(line)
            nn += 1
        elif (line =='***\n'):
            ii = 0
            counter += 1
            nn = 0
            qlist.append(chunks)
            qlistInit.append(chunksInit)
            qlistFinal.append(chunksFinal)
            chunks = []
            chunksInit = []
            chunksFinal = []
        else:
            nn += 1
    return qlistInit, qlist, qlistFinal

def clearCells(workbook, cells):
    '''
    clear values in workbook idicated by cells
    example of how to input cells --> for row in ws['A1:G37'] 
    '''     
    for row in workbook[cells]:
        for cell in row:
            cell.value = None


#-------------------------------------------------------------------------------------
# USER INPUTS
# select file name to read (name only)
meaFile = 'Conimed_Meatest RP-100_202104_all'
language = 'ENG'  # 'ENG' or 'ESP', this will impact on result because 
xlsx_name = 'protocolo_RM_rev1.1' # protocol to be used
n_estadistica_medicion = 20

#-------------------------------------------------------------------------------------
#   Save data to xlsx file 

mea_file_name = meaFile + '.mea'
lines = openFile(mea_file_name) # open file and read lines
measurement_repetitions = lines.count(lines[0])
print('Número de mediciones totales = {}'.format(measurement_repetitions))
    
# read mea_file_name and save meas to qlist, it has a different file in each list element qlist[i]
qlistInit, qlist, qlistFinal = read_mea_file(lines)

# excel changes equations depending on language, so I have two different protocols depending on language
xlsx_extension = '.xlsx'
xlsx_file = xlsx_name + '_' + language + xlsx_extension

# Load workbook and copy nn-times the protocol sheet (0), nn= number of full measurements.
xl = oxl.load_workbook(xlsx_file)
target = xl['(0)']

for nn in range(1,measurement_repetitions):
    copied = xl.copy_worksheet(target)
    copied.title = '({sheet_number})'.format(sheet_number= nn) # assign name of new sheet

# save the changes to protocol, close the original and reload:
xl.save('working_protocol.xlsx')
xl.close()
xlsx_file = 'working_protocol.xlsx'
xl = oxl.load_workbook(xlsx_file) 
print('MAR line {}: sumar en xl.save() un check para ver si el archivo ya exite. Esta función realiza el save y sobreescribe sin avisar.'.format(getframeinfo(currentframe()).lineno))


# sheet number sheetZero should be sheet named '(0)'
n_sheetZero = 4
sheet_name = xl.sheetnames[n_sheetZero]

# check if the sheets are in the right position, note that it depends on the order of the sheets that's why I include an if before the for, so it rises an exception if it is different from expected
if (sheet_name != '(0)'):
    cf = currentframe()
    raise Exception('File: {filename}, error in line{linenumber}: The xlsx shets have been moved, please check that you are using the original \' {protocolfile} \' file'.format(filename= getframeinfo(cf).filename, linenumber= cf.f_lineno, protocolfile= xlsx_name))

#       0    1   2   3   4   5   6   7   8   9
col1 = ['C','D','E','G','H','I','J','L','O','P'] # columns where the data of qlist should be saved
row1 = 83 # first row where the data of qlist should be saved
for qq in range(0,len(qlist)):
    # activate sheet, skips first n_sheetZero sheets and starts at sheet (0)
    sheet = xl[xl.sheetnames[n_sheetZero + qq]]

    # save data of qlist into the sheets
    ii = 0
    for dato in qlist[qq]:
        datosplit = dato.split(';')
        row2 = str(row1 + ii)
        sheet['B' + row2] = float(ii) # save meas number
        sheet[col1[0] + row2] = float(datosplit[0])
        sheet[col1[1] + row2] = float(datosplit[1])
        sheet[col1[2] + row2] = datosplit[2]
        # formulas to include
        sheet[col1[3] + row2] = "=$P$9*(1+$U$9*($R$15-20)+$V$9*($R$15-20)^2)"
        sheet[col1[4] + row2] = "="+ col1[0]+ str(row2)+ "*"+ col1[3]+ row2 
        sheet[col1[5] + row2] = "=("+col1[4]+str(row2)+"-$P$6)"
        sheet[col1[6] + row2] = "=("+col1[4]+str(row2)+"-$P$6)*1000000/$P$6)"

        if language == 'ENG':
            sheet[col1[7] + row2] = "=_xlfn.AVERAGE(" + col1[4] + str(row1) + ":" + col1[4] + row2 + ")"
            sheet[col1[8] + row2] = "=_xlfn.STDEV.P(" + col1[4] + str(row1) + ":" + col1[4] + row2 + ")"
            # its necessary to use the _xlfn only in the STDEV function, the rest will work fine. This will avoid excell incluiding an @ in front of the equations.
        elif language == 'ESP':
            sheet[col1[7] + row2] = "=_xlfn.PROMEDIO(" + col1[4] + str(row1) + ":" + col1[4] + row2 + ")"
            sheet[col1[8] + row2] = "=_xlfn.DESVEST.P(" + col1[4] + str(row1) + ":" + col1[4] + row2 + ")"
        else:
            try:
                raise ValueError()
            except:
                print(traceback.format_exc())
                print('Languaje wrong, only accepting ENG or ESP')
        ii +=1
    
    str_range =  col1[4] + str(len(qlist[qq]) -1 + row1 - n_estadistica_medicion) + ":" + col1[4] + str(len(qlist[qq]) -1 + row1) 
    if language == 'ENG':
        sheet[col1[8] + str(23)] = "=_xlfn.AVERAGE(" + str_range + ")" 
        sheet[col1[9] + str(23)] = "=_xlfn.STDEV.P(" + str_range + ")*1e6/O23" 
    elif language == 'ESP':
        sheet[col1[8] + str(23)] = "=_xlfn.PROMEDIO(" + str_range + ")"
        sheet[col1[9] + str(23)] = "=_xlfn.DESVEST.P(" + str_range + ")*1e6/O23" 
    
    # remove cells that are not in the measured range (they are preloaded in the xlsx protocol file)
    celdas_borrar_init = len(qlist[qq]) + row1
    celdas_borrar_final = 500 # I assume this is the last cell with values, if changed in the prtotocol this must be modify acordingly.
    celdas_str = 'A{initial}:O{final}'.format(initial=celdas_borrar_init, final=celdas_borrar_final)
    clearCells(sheet, celdas_str)
  
    # save data of qlistInit into the sheets
    jj = 1
    for datoInit in qlistInit[qq]:
        sheet[col1[0] + str(jj)] = datoInit.rstrip()
        jj += 1
    jj = 3
    for datoFinal in qlistFinal[qq]:
        sheet['F' + str(jj)] = datoFinal.rstrip()
        jj += 1

# Save data changes to new excel file
xl.save(meaFile + '_meas' + xlsx_extension)
if os.path.exists('working_protocol.xlsx'):
  os.remove('working_protocol.xlsx')

print('The file has been saved as: {}'.format(os.getcwd() + '\\' + meaFile + '_meas' + xlsx_extension))
